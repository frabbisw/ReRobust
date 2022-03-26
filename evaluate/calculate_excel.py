import openpyxl as opx


def calculate_reduce():
    file_name = 'experiment result.xlsx'
    sheet_name = 'sheet1'
    workbook = opx.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.insert_cols(idx=6)
    for index in range(1, sheet.max_row + 1):
        if (index % 2) != 0:
            continue
        Fi = 'F' + str(index)
        Ei_1 = 'E' + str(index - 1)
        Ei = 'E' + str(index)
        FORMULA = '=' + Ei_1 + '-' + Ei
        print(FORMULA)
        sheet[Fi] = FORMULA
    workbook.save(file_name)
    workbook.close()


def calculate_percentage():
    file_name = 'experiment result.xlsx'
    sheet_name = 'sheet1'
    workbook = opx.load_workbook(file_name)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.insert_cols(idx=7)
    for index in range(1, sheet.max_row + 1):
        if (index % 2) != 0:
            continue
        Gi = 'G' + str(index)
        Ei_1 = 'E' + str(index - 1)
        Ei = 'E' + str(index)
        # =ROUND((E5-E6)/E5*100,2)
        FORMULA = '=ROUND((' + Ei_1 + '-' + Ei + ')/' + Ei_1 + '*100,2)'
        print(FORMULA)
        sheet[Gi] = FORMULA
    workbook.save(file_name)
    workbook.close()


def performance():
    file_name = 'experiment result.xlsx'
    sheet_name = 'sheet1'
    workbook = opx.load_workbook(file_name, data_only=True)
    sheet = workbook.get_sheet_by_name(sheet_name)
    sheet.insert_cols(idx=7)
    for index in range(1, sheet.max_row + 1):
        if (index % 2) != 0:
            continue
        Fi = 'F' + str(index)
        Gi = 'G' + str(index)

        a1 = '33.33%'
        a1_int = percent_to_int(a1)
        a2 = '66.66%'
        a2_int = percent_to_int(a2)
        a3 = '100%'
        a3_int = percent_to_int(a3)
        print(sheet[Fi].value)
        Fi_int = percent_to_int(sheet[Fi].value)

        if Fi_int < a1_int:
            sheet[Gi] = '☆'
        elif Fi_int < a2_int:
            sheet[Gi] = '☆☆'
        else:
            sheet[Gi] = '☆☆☆'
    workbook.save(file_name)
    workbook.close()


def percent_to_int(string):
    if "%" in string:
        newint = float(string.strip("%")) / 100
        return newint
    else:
        print("你输入的不是百分比！")


if __name__ == '__main__':
    calculate_reduce()
    calculate_percentage()
    # performance()
